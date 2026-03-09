from __future__ import annotations

from datetime import date

from .calendar_utils import month_range
from .domain import AlertItem, Assignment, MonthInput, SHIFT_HOURS


def _slot_key_to_code(slot_key: str) -> str:
    return slot_key.split("#", 1)[0]


def build_staff_day_codes(mi: MonthInput, assignments: tuple[Assignment, ...]) -> dict[tuple[str, date], str]:
    out: dict[tuple[str, date], str] = {}
    for a in assignments:
        for slot_key, sid in a.slots.items():
            out[(sid, a.day)] = _slot_key_to_code(slot_key)
    return out


def compute_summary(
    mi: MonthInput, assignments: tuple[Assignment, ...]
) -> list[tuple[str, str, int, float, int]]:
    start, end = month_range(mi.month)
    total_days = end.day
    by_day = build_staff_day_codes(mi, assignments)

    rows: list[tuple[str, str, int, float, int]] = []
    for s in mi.staff:
        work_days = 0
        work_hours = 0.0
        for d in range(1, total_days + 1):
            cur = date(start.year, start.month, d)
            code = by_day.get((s.id, cur))
            if not code:
                continue
            work_days += 1
            work_hours += SHIFT_HOURS.get(code, 0.0)
        rest_days = total_days - work_days
        rows.append((s.id, s.name, work_days, work_hours, rest_days))
    return rows


def export_xlsx(
    mi: MonthInput,
    assignments: tuple[Assignment, ...],
    out_path: str,
    alerts: tuple[AlertItem, ...] = (),
    request_violations: tuple[tuple[date, str], ...] = (),
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as e:
        raise RuntimeError(
            "openpyxl が見つかりません。`pip install -r requirements.txt` を実行してください。"
        ) from e

    wb = Workbook()
    ws = wb.active
    ws.title = mi.month

    fill_header = PatternFill("solid", fgColor="1F2937")
    font_header = Font(color="FFFFFF", bold=True)

    start, end = month_range(mi.month)
    total_days = end.day
    weekdays = "月火水木金土日"

    header = ["名前"] + [f"{d}({weekdays[date(start.year, start.month, d).weekday()]})" for d in range(1, total_days + 1)] + [
        "勤務日数",
        "勤務時間",
        "休日数",
    ]
    ws.append(header)

    for col in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    by_day = build_staff_day_codes(mi, assignments)
    summaries = compute_summary(mi, assignments)

    for sid, name, work_days, work_hours, rest_days in summaries:
        row = [name]
        for d in range(1, total_days + 1):
            cur = date(start.year, start.month, d)
            row.append(by_day.get((sid, cur), ""))
        row.extend([work_days, work_hours, rest_days])
        ws.append(row)

    ws.freeze_panes = "B2"
    ws.column_dimensions["A"].width = 16
    for col in range(2, 2 + total_days):
        ws.column_dimensions[get_column_letter(col)].width = 8
    ws.column_dimensions[get_column_letter(2 + total_days)].width = 10
    ws.column_dimensions[get_column_letter(3 + total_days)].width = 10
    ws.column_dimensions[get_column_letter(4 + total_days)].width = 10

    for r in range(2, 2 + len(summaries)):
        for c in range(1, len(header) + 1):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")

    ws2 = wb.create_sheet("不足アラート")
    ws2_header = ["日付", "不足シフト", "不足人数"]
    ws2.append(ws2_header)
    for col in range(1, len(ws2_header) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for a in alerts:
        ws2.append([a.date.isoformat(), a.shift_code, a.missing_count])
    for d, sid in request_violations:
        ws2.append([d.isoformat(), f"希望休違反({sid})", 1])

    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 10
    for r in range(2, 2 + len(alerts) + len(request_violations)):
        for c in range(1, 4):
            ws2.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")

    wb.save(out_path)
